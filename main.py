from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment
import json
import customtkinter as CTk
import tkinter as tk
from PIL import Image
from pathlib import Path
import os
import sys

#Build command:
#pyinstaller --onefile --windowed --add-data "close.png:." --add-data "settings.png:." --add-data "folder.png:." .\main.py

#Сustom class for the drag and drop system which is based on the base listBox class in Tkinter
class DraggableListbox(CTk.CTkFrame):
    def __init__(self, master, items, **kwargs):
        super().__init__(master, **kwargs)
        self.listbox = tk.Listbox(self, font=("Inter", 16), width=40)
        self.listbox.pack(fill=tk.BOTH, expand=True)
        #parsing the array that came as a parameter and adding it to the listBox
        for item in items:
            self.listbox.insert(tk.END, item)

        #Bind to select an element to drag
        self.listbox.bind("<Button-1>", self.save_mouse_pos)

        #Bind for dragging
        self.listbox.bind("<B1-Motion>", self.on_drag)

        #Variable for current dragging item
        self.selected_index = None

    #When the left mouse button is pressed it search for the closest element to the cursor and save it
    def save_mouse_pos(self, event):
        self.selected_index = self.listbox.nearest(event.y)

    #When dragging swaps elements
    def on_drag(self, event):
        #Get selected item from listbox
        selected_item = self.listbox.get(self.selected_index)
        #Find nearest index to current mouse position
        current_index = self.listbox.nearest(event.y)

        #If new index is different from selected index
        if current_index != self.selected_index:
            #Remove item from old position
            self.listbox.delete(self.selected_index)
            #Insert item to new position
            self.listbox.insert(current_index, selected_item)
            #Update selected index
            self.selected_index = current_index

    #Return elements in the order set by user
    def get_items(self):
        return self.listbox.get(0, tk.END)

#Class for working with files
class FileManager:
    def __init__(self, config_file = "config.json"):
        self.config_file = config_file
        self.config = self._load_config()

    #Load a config file
    def _load_config(self):
        with open(self.config_file, 'r') as file:
            return json.loads(file.read())

    #Wite new config to config file
    def write_config(self, config_manager):
        #New config
        config = {}
        #Iterate over the attributes of the config class
        for attr in dir(config_manager):
            #Weed out the basic methods
            if not attr.startswith('__'):
                #Get value from config attribute
                value = getattr(config_manager, attr)
                #Write it to new config
                config[attr] = value
        #Write new config to config file
        with open(self.config_file, "w") as file:
            file.write(json.dumps(config))

#Class for config
class ConfigManager:
    def __init__(self, config):
            #Get required fields
            self.templates = config["templates"]
            self.save_folder_path = config["save_folder_path"]

#Class parse html and extract the needed information
class HtmlParcer:
    def __init__(self, html):
        #Parse html
        self.soup = BeautifulSoup(html, "html.parser")
        self.fields_list = self._parse_fields()
        self.progress_list = self._parse_progress()

    #Function return a list of fields
    def _parse_fields(self):
        #Search th tags in the thead, there are located fields
        th_tags =  self.soup.thead.find_all("th", class_='center')
        #Array for fields, names of students is required, so we create it immediately.
        list = ["Opiskelijan nimi"]

        #Loop iterate through th tags
        for th_tag in th_tags:
            #Get the data-tooltip parameter which contains details about field
            data_tooltip = th_tag.get('data-tooltip')
            #Remove unnecessary symbols
            data_dict = json.loads(data_tooltip.replace('&quot;', ''))
            #Extract the parameter in which the name is located
            fields_name = data_dict.get('Opintojakson/tutkinnon osan nimi')
            #Check if the field is required or optional and write it
            if data_dict.get("Koodi"):
                if data_dict.get("Koodi") == "pak":
                    fields_name = f"Pak.{data_dict.get('Opintojakson/tutkinnon osan nimi')}"
                else:
                    fields_name = f"Val.{data_dict.get('Opintojakson/tutkinnon osan nimi')}"
            #add field to list
            list.append(fields_name)
        return(list)

    #parse merits students, they are all found in tr tag
    def _parse_progress(self):
        return self.soup.tbody.find_all("tr")

#Data processing takes place in this class
class DataManager:
    def __init__(self, app):
        self.checkboxes_list = app.gui.checkboxes_list
        self.entry_list = app.gui.entry_list
        self.parser = app.html_parser
        self.config = app.config_manager
        self.selected_fields = self._initialize_selected_fields()
    
    #Prepares data for _process_data function
    def process_data(self, new_selected_fields = None):
        #If nothing was selected then stop function
        if(new_selected_fields):
            self.selected_fields = new_selected_fields
        self.table = self._initialize_table()
        self.indexes_list = self._initialize_indexes()
        self._process_data()

    #Initializes an array of selected industries, which contains other arrays
    #First element is the actual name of the field and the second is the desired name
    def _initialize_selected_fields(self):
        #Names of the students must be, so we create them
        fields_list = [["Opiskelijan nimi"]]

        #loop iterate through all the checkboxes
        for n,i in enumerate(self.checkboxes_list):
            #Check if checkbox selected
            if i.get() == 1:
                #get checkbox text
                element = [i.cget("text")]
                #check if input not empty
                if len(self.entry_list[n].get()) != 0:
                    #get text from input
                    element.append(self.entry_list[n].get())
                #append element to fields list array
                fields_list.append(element)
        return fields_list
    #Initialization of the table that will later be filled and written in excel
    def _initialize_table(self):
        table = {}
        #Students name is always required
        table["Opiskelijan nimi"] = []
        #Creating empty fields in a table
        for i in self.selected_fields:
            table[i[-1]] = []
        return table

    #initialization of the list of indexes of the fields that we need
    def _initialize_indexes(self):
        indexes_list = []
        for i in self.selected_fields:
            if i[0] in self.parser.fields_list:
                indexes_list.append(self.parser.fields_list.index(i[0]))
        return indexes_list

    #Function delete not needed spaces from string
    def _delete_spaces(self, text):
        return ''.join(char for char in text if char.isalnum())

    #Main function that write data in a table
    def _process_data(self):
        #Loop iterate students progress list
        for n,row in enumerate(self.parser.progress_list):
            #Skip first element
            if n == 0:
                continue
            #Get all row
            elements = row.find_all("td")
            #Loop iterate indexes list
            for idx, selected_index in enumerate(self.indexes_list):
                field_name = self.selected_fields[idx][-1]
                element = elements[selected_index]
                #Names and metrits are in different tags, so need to check where to pull the text from
                value = element.a.text if element.find("a") else self._delete_spaces(element.text)
                #Replase o to X
                value = "X" if value.lower() == "o" else value
                self.table[field_name].append(value)

        #Loop is combine fields with the same name
        for i in self.table:
            #Code above adds the same data with the same name to the same array,
            #so we may end up with all arrays having a length of 12 and one having a length of 24 or 36
            #Check with the name array because it is always same
            #We check the length of the array of names, so it is always the correct length
            if len(self.table[i]) % len(self.table["Opiskelijan nimi"]) == 0 and len(self.table[i]) != len(self.table["Opiskelijan nimi"]):
                list = self.table[i]
                #Variable where we write how many times the array is larger than we need
                step = int(len(list) / len(self.table["Opiskelijan nimi"]))
                #New array that will be shorter by reducing it by 'step'
                new_array=[]
                #Loop iterate list by a step
                for j in range(0, len(list),step):
                    #Get a slice of the list based on the step
                    current_elements = list[j:j + step]
                    new_element = ""
                    #Add up the elements in the current slice, skipping empty values
                    for k in current_elements:
                        if k != "":
                            if new_element == "":
                                new_element = int(k)
                            else:
                                new_element = new_element + int(k)
                    #Add the summed up value to the new array
                    new_array.append(new_element)
                #Replace the old array with the new one
                self.table[i] = new_array

#Class for writing data in Excel and styling
class ExcelWriter:
    def __init__(self, table, total_lines, folder_path, filename="students"):
        self.table = table
        self.df = pd.DataFrame(table)
        self.total_lines = total_lines
        self.folder_path = folder_path
        self.filename = filename
        #if need, add empty rows
        if  len(self.df) < self.total_lines:
            #Number of required empty rows to be added
            empty_rows = self.total_lines - len(self.df)
            #Array of empty elements
            empty_data = pd.DataFrame([[""] * len(self.df.columns)] * empty_rows, columns=self.df.columns)
            #Combine dataframes
            self.df = pd.concat([self.df, empty_data], ignore_index=True)
        self._write_to_excel()

    #Function write the table to an excel file.
    def _write_to_excel(self):
        #Path to the folder to save to
        folder_path = ""
        #If folder_path not empty
        if self.folder_path:
            #Add / because its path
            folder_path = self.folder_path+"/"
        #Save and open excel file for styling
        with pd.ExcelWriter(f"{folder_path}{self.filename}.xlsx", engine='openpyxl') as writer:
            #Get sheets from table
            self.df.to_excel(writer, sheet_name='Table', index=False)
            #Get current sheet from table
            work_sheet = writer.sheets['Table']
            self._adjust_columns(work_sheet)
            self._apply_styles(work_sheet)

    #Function sets column the width, which depends on the maximum width
    def _adjust_columns(self, sheet):
        #Min width of column
        base_width = 10
        #List for max widths of columns
        column_widths = {}
        #Loop iterate columns and put column width to list
        for column in sheet.columns:
            max_length = 0
            #Get the letter of the current column
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    #If the cell text is longer than current max, update it
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            #Save the max length for the column
            column_widths[column_letter] = max_length
        #set width for A
        if 'A' in column_widths:
            sheet.column_dimensions['A'].width = column_widths['A']

        #Find the max width for all columns except 'A'
        max_width_for_others = max(width for letter, width in column_widths.items() if letter != 'A')
        #Set max width for other
        for column_letter in column_widths:
            if column_letter != 'A':
                sheet.column_dimensions[column_letter].width = max(max_width_for_others, base_width)

    #Add styles to the table, such as border and fill
    def _apply_styles(self, sheet):
        #Gray color fill
        fill = PatternFill(start_color="828181", end_color="828181", fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        #Add border for the cell and fill the rows with a step with gray color
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = center_alignment
                if cell.row % 2 == 0:
                    cell.fill = fill
#Class for working with graphics
class Gui(CTk.CTk): #TODO описать класс
    def __init__(self, app, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self._app = app
        
        #Constants are placed in one place for convenience
        self._FONT = "Inter"
        self._WHITE_COLOR = "#EBEBEB"
        self._BLUE_COLOR = "#1256D3"
        self._PURPLE_COLOR = "#5A4AEA"
        self._HOVER_PURPLE_COLOR = "#4533b5"

        # self._NEXT_BUTTON_TEXT = "Next"
        # self._BACK_BUTTON_TEXT = "Back"
        # self._SAVE_BUTTON_TEXT = "Save"
        # self._EXIT_BUTTON_TEXT = "Exit"
        # self._MENU_BUTTON_TEXT = "Menu"

        # self._HEADER_TEXT = "Teacher helper"
        # self._MENU_COPY_AND_CLICK_TEXT = "Copy code, and click the button"
        # self._MENU_BUTTON_TEXT = "Paste"
        # self._MENU_ERR_WRONG_HTML_CODE_TEXT = "Wrong html code"
        # self._MENU_ERR_CANT_GET_CLIPBOARD_TEXT = "Cant get data from clipboard"
        
        # self._SETTINGS_TITLE_TEXT = "Settings"
        # self._SETTINGS_SELECT_FOLDER_TEXT = "Select a folder to save the files:"
        # self._SETTINGS_DELETE_TEMPLATES_TEXT = "Delete template"
        # self._SETTINGS_NOT_FOUND_TEXT = "Not found"

        # self._MAIN_TITLE_TEXT = "Choose template"
        # self._MAIN_ALL_BUTTON_TEXT = "All"
        # self._MAIN_CLEAR_BUTTON_TEXT = "Clear"
        # self._MAIN_WRITE_FILE_NAME_TEXT = "Write file name:"
        # self._MAIN_WRITE_LINES_COUNT_TEXT = "Write the numbe of lines:"
        # self._CUSTOM_ORDER_TEXT = "Custom order"

        # self._MAIN_MODAL_WINDOW_TITLE = "Write template name"

        # self._CUSTOM_ORDER_TITLE = "Custom order:"

        # self._SUCCESS_TITLE = "Success!"

        self._NEXT_BUTTON_TEXT = "Seuraava"
        self._BACK_BUTTON_TEXT = "Takaisin"
        self._SAVE_BUTTON_TEXT = "Tallenna"
        self._EXIT_BUTTON_TEXT = "Poistu"
        self._MENU_BUTTON_TEXT = "Valikko"

        self._HEADER_TEXT = "Opettajan apuri"
        self._MENU_COPY_AND_CLICK_TEXT = "Kopioi koodi ja klikkaa painiketta"
        self._MENU_PASTE_BUTTON_TEXT = "Liitä"
        self._MENU_ERR_WRONG_HTML_CODE_TEXT = "Virheellinen HTML-koodi"
        self._MENU_ERR_CANT_GET_CLIPBOARD_TEXT = "Ei voi saada tietoja leikepöydältä"

        self._SETTINGS_TITLE_TEXT = "Asetukset"
        self._SETTINGS_SELECT_FOLDER_TEXT = "Valitse kansio tiedostojen tallentamista varten:"
        self._SETTINGS_DELETE_TEMPLATES_TEXT = "Poista malli"
        self._SETTINGS_NOT_FOUND_TEXT = "Ei löydy"

        self._MAIN_TITLE_TEXT = "Valitse malli"
        self._MAIN_ALL_BUTTON_TEXT = "Kaikki"
        self._MAIN_CLEAR_BUTTON_TEXT = "Tyhjennä"
        self._MAIN_WRITE_FILE_NAME_TEXT = "Kirjoita tiedoston nimi:"
        self._MAIN_WRITE_LINES_COUNT_TEXT = "Kirjoita rivien määrä:"
        self._CUSTOM_ORDER_TEXT = "Mukautettu järjestys"

        self._MAIN_MODAL_WINDOW_TITLE = "Kirjoita mallin nimi"

        self._CUSTOM_ORDER_TITLE = "Mukautettu järjestys:"

        self._SUCCESS_TITLE = "Onnistui!"

        #Basic configuration of tkinter
        self.geometry("800x500")
        self.title("Teacher helper")
        self._set_appearance_mode("light")
        self.resizable(height=False, width=False)

        #Load images
        self._settings_image = CTk.CTkImage(light_image=Image.open(self._app.resource_path("settings.png")), dark_image=Image.open(self._app.resource_path("settings.png")), size=(50,50))
        self._folder_image = CTk.CTkImage(light_image=Image.open(self._app.resource_path("folder.png")), dark_image=Image.open(self._app.resource_path("folder.png")), size=(20,20))
        self._delete_image = CTk.CTkImage(light_image=Image.open(self._app.resource_path("close.png")), dark_image=Image.open(self._app.resource_path("close.png")), size=(15,15))

        #Load first window
        self.load_menu()

    def load_menu(self):
        self._clear()
        self._load_header(is_show_settings=True)
        menu_text = CTk.CTkLabel(master=self,fg_color=self._WHITE_COLOR,text_color="black",text=self._MENU_COPY_AND_CLICK_TEXT, font=(self._FONT,32))
        menu_text.grid(row=1, column=0, pady=(90, 0))

        #Label for errors
        self.error_text = CTk.CTkLabel(master=self,fg_color=self._WHITE_COLOR,text_color="red",text="", font=(self._FONT,24))
        self.error_text.grid(row=2, column=0, pady=(5))

        menu_button = CTk.CTkButton(master=self, command=self._app.menu_button_handle,hover_color=self._HOVER_PURPLE_COLOR,width=170,height=45, text_color="black",corner_radius=11,border_width=1,border_color="black",text=self._MENU_PASTE_BUTTON_TEXT,font=(self._FONT,30), fg_color=self._PURPLE_COLOR, bg_color=self._WHITE_COLOR)
        menu_button.grid(row=3, column=0, pady=(10,0))

    def load_main(self):
        self._clear()
        self._load_header()
        self._app.filename = ""

        BUTTON_PADDING = 5
        BUTTON_FONT_SIZE = 18

        hero_text = CTk.CTkLabel(master=self, fg_color=self._WHITE_COLOR, text_color="black", text=self._MAIN_TITLE_TEXT, font=(self._FONT, 24))
        hero_text.grid(row=1,column=0, sticky="w", padx=(7,0), pady=(7,0))

        template_buttons_frame = CTk.CTkFrame(master=self, fg_color=self._WHITE_COLOR, bg_color=self._WHITE_COLOR)
        template_buttons_frame.grid(row=2, column=0, sticky="w")

        #Clear all selected fields
        template_clear = CTk.CTkButton(master=template_buttons_frame,command=lambda: self._app.select_all_checkboxes(0),hover_color=self._HOVER_PURPLE_COLOR,width=80,height=20, text_color="black",corner_radius=11,border_width=1,border_color="black",text=self._MAIN_CLEAR_BUTTON_TEXT,font=(self._FONT,BUTTON_FONT_SIZE), fg_color=self._PURPLE_COLOR, bg_color=self._WHITE_COLOR)
        template_clear.grid(row=0, column=0, sticky="w", padx=(BUTTON_PADDING,0),pady=(BUTTON_PADDING, 0))

        #Select all fields
        template_all = CTk.CTkButton(master=template_buttons_frame, command=lambda: self._app.select_all_checkboxes(1),hover_color=self._HOVER_PURPLE_COLOR,width=80,height=20, text_color="black",corner_radius=11,border_width=1,border_color="black",text=self._MAIN_ALL_BUTTON_TEXT,font=(self._FONT,BUTTON_FONT_SIZE), fg_color=self._PURPLE_COLOR, bg_color=self._WHITE_COLOR)
        template_all.grid(row=0, column=1,sticky="w", padx=(BUTTON_PADDING,0),pady=(BUTTON_PADDING, 0))

        #Insert templates
        for n,i in enumerate(self._app.config_manager.templates):
            button = CTk.CTkButton(master=template_buttons_frame,command=lambda name=i: self._app.select_checkboxes_by_template(name),hover_color=self._HOVER_PURPLE_COLOR,width=80,height=20, text_color="black",corner_radius=11,border_width=1,border_color="black",text=i,font=(self._FONT,BUTTON_FONT_SIZE), fg_color=self._PURPLE_COLOR, bg_color=self._WHITE_COLOR)
            button.grid(row=0, column=2+n,sticky="w", padx=(BUTTON_PADDING,0),pady=(BUTTON_PADDING, 0))

        #Insert add template button, can be maximum 3 templates
        if len(self._app.config_manager.templates) < 3:
            template_add = CTk.CTkButton(master=template_buttons_frame,command=self.load_add_tamplate_modal_window,hover_color=self._HOVER_PURPLE_COLOR,width=30,height=20, text_color="black",corner_radius=11,border_width=1,border_color="black",text="+",font=(self._FONT,BUTTON_FONT_SIZE), fg_color=self._PURPLE_COLOR, bg_color=self._WHITE_COLOR)
            template_add.grid(row=0, column=2+len(self._app.config_manager.templates),sticky="w",padx=(BUTTON_PADDING,0),pady=(BUTTON_PADDING,0))

        checkbox_frame = CTk.CTkScrollableFrame(master=self,fg_color="#D5D5D5",bg_color=self._WHITE_COLOR, width=380, height=300)
        checkbox_frame.grid(row=3, column=0, sticky="w", padx=(10,0), pady=(10,0))

        self.checkboxes_list = []
        self.entry_list = []

        #Insert fields checkboxes from html
        for n,i in enumerate(self._app.html_parser.fields_list):
            if n == 0:
                continue

            state = "disabled"

            checkbox_var = CTk.IntVar(value=0)
            checkbox_var.trace_add("write", lambda var_name, index, mode, value=n: self._app.on_select_checkbox(var_name, index, mode, value))
            checkbox = CTk.CTkCheckBox(master=checkbox_frame,variable=checkbox_var, text=i, font=(self._FONT,14), text_color="black", checkbox_width=20, checkbox_height=20)
            checkbox.grid(row=n*2, column=0, sticky="w")
            self.checkboxes_list.append(checkbox)

            entry = CTk.CTkEntry(master=checkbox_frame, state=state,font=(self._FONT, 16),width=200, fg_color="#D5D5D5",text_color="black")
            entry.grid(row=n*2+1, column=0, sticky="w",pady=(0,10), padx=(25,0))
            entry.bind("<Return>", self._app.focus_next_entry)
            self.entry_list.append(entry)

        additions_frame = CTk.CTkFrame(master=self,bg_color=self._WHITE_COLOR, fg_color=self._WHITE_COLOR)
        additions_frame.grid(row=3, column=0, sticky="ne", padx=(0,40))

        file_name_frame_text = CTk.CTkLabel(master=additions_frame, text=self._MAIN_WRITE_FILE_NAME_TEXT, font=(self._FONT, 24),text_color="black")
        file_name_frame_text.grid(row=0,column=0)

        #Input for file name
        self.file_name_frame_entry = CTk.CTkEntry(master=additions_frame, text_color="black",fg_color=self._WHITE_COLOR, font=(self._FONT,18))
        self.file_name_frame_entry.grid(row=1,column=0, pady=(10,0))

        column_count_frame_text = CTk.CTkLabel(master=additions_frame, text=self._MAIN_WRITE_LINES_COUNT_TEXT, font=(self._FONT, 24),text_color="black")
        column_count_frame_text.grid(row=2,column=0, pady=(30,0))

        #Input for rows count
        self.column_count_frame_entry = CTk.CTkEntry(master=additions_frame, text_color="black",fg_color=self._WHITE_COLOR, font=(self._FONT,18))
        self.column_count_frame_entry.insert(0, 0)
        self.column_count_frame_entry.grid(row=3,column=0, pady=(10,0))

        self.is_custom_order = CTk.IntVar(value=0)
        checkbox_is_custom_order = CTk.CTkCheckBox(master=additions_frame,bg_color=self._WHITE_COLOR,variable=self.is_custom_order, text=self._CUSTOM_ORDER_TEXT, font=(self._FONT,22), text_color="black", checkbox_width=20, checkbox_height=20)
        checkbox_is_custom_order.grid(row=4, column=0, pady=(30,0))

        back_button = CTk.CTkButton(master=self,command=lambda: self._app.change_window(0),hover_color=self._HOVER_PURPLE_COLOR, text=self._BACK_BUTTON_TEXT, fg_color=self._PURPLE_COLOR, font=(self._FONT, 18), bg_color=self._WHITE_COLOR, width=70, border_width=1, border_color="black", text_color="black")
        back_button.grid(row=4, column=0, sticky="ws", padx=(10,0), pady=(5,0))

        next_button = CTk.CTkButton(master=self,command=self._app.compilate_data,hover_color=self._HOVER_PURPLE_COLOR, text=self._NEXT_BUTTON_TEXT, fg_color=self._PURPLE_COLOR, font=(self._FONT, 18), bg_color=self._WHITE_COLOR, width=70, border_width=1, border_color="black", text_color="black")
        next_button.grid(row=4, column=0, sticky="se", padx=(0,10), pady=(5,0))

    def load_success(self):
        self._clear()
        self._load_header()

        success_text = CTk.CTkLabel(master=self, text=self._SUCCESS_TITLE, font=(self._FONT,32),fg_color=self._WHITE_COLOR,text_color="black")
        success_text.grid(row=1, column=0, pady=(100, 0))

        success_frame = CTk.CTkFrame(master=self, bg_color=self._WHITE_COLOR, fg_color=self._WHITE_COLOR)
        success_frame.grid(row=2,column=0, pady=(15,0))

        exit_button = CTk.CTkButton(master=success_frame,command=self.quit,bg_color=self._WHITE_COLOR,width=100, text=self._EXIT_BUTTON_TEXT, font=(self._FONT,24), fg_color=self._PURPLE_COLOR,hover_color=self._HOVER_PURPLE_COLOR, border_width=1, border_color="black",text_color="black")
        exit_button.grid(row=0,column=0)

        exit_button = CTk.CTkButton(master=success_frame,command=lambda:self._app.change_window(1),bg_color=self._WHITE_COLOR,width=100, text=self._BACK_BUTTON_TEXT, font=(self._FONT,24), fg_color=self._PURPLE_COLOR,hover_color=self._HOVER_PURPLE_COLOR, border_width=1, border_color="black",text_color="black")
        exit_button.grid(row=0,column=1, padx=10)

        exit_button = CTk.CTkButton(master=success_frame,command=lambda:self._app.change_window(0),bg_color=self._WHITE_COLOR,width=100, text=self._MENU_BUTTON_TEXT, font=(self._FONT,24), fg_color=self._PURPLE_COLOR,hover_color=self._HOVER_PURPLE_COLOR, border_width=1, border_color="black",text_color="black")
        exit_button.grid(row=0,column=2)

        author_text = CTk.CTkLabel(master=self,font=(self._FONT, 16),text_color="black",bg_color=self._WHITE_COLOR, fg_color=self._WHITE_COLOR, text="Author: Huziichuk Nazar | Github: guziiuchyk/Teacher-helper | Gmail: guziiuchyk@gmail.com")
        author_text.grid(row=3, column=0, sticky="s",pady=(210,0))

    def load_custom_order(self):
        self._clear()
        self._load_header()

        custom_order_text = CTk.CTkLabel(master=self, text=self._CUSTOM_ORDER_TITLE, font=(self._FONT,30),fg_color=self._WHITE_COLOR,text_color="black")
        custom_order_text.grid(row=1, column=0, pady=(10,0))

        self.draggable_listbox = DraggableListbox(self, self._app.get_select_fields_for_drag())
        self.draggable_listbox.grid(row=2, column=0,pady=10)

        next_button = CTk.CTkButton(master=self,command=lambda: self._app.change_window(1),hover_color=self._HOVER_PURPLE_COLOR, text=self._BACK_BUTTON_TEXT, fg_color=self._PURPLE_COLOR, font=(self._FONT, 18), bg_color=self._WHITE_COLOR, width=70, border_width=1, border_color="black", text_color="black")
        next_button.grid(row=3, column=0, sticky="sw", padx=(10,0), pady=(125,0))

        back_button = CTk.CTkButton(master=self,command=lambda: self._app.compilate_ordered_data(self.draggable_listbox.get_items()),hover_color=self._HOVER_PURPLE_COLOR, text=self._NEXT_BUTTON_TEXT, fg_color=self._PURPLE_COLOR, font=(self._FONT, 18), bg_color=self._WHITE_COLOR, width=70, border_width=1, border_color="black", text_color="black")
        back_button.grid(row=3, column=0, sticky="se", padx=(0,10), pady=(125,0))

    def load_settings(self):
        self._clear()
        self._load_header()

        GRAY_COLOR = "#c9c9c9"

        frame = CTk.CTkFrame(master=self, corner_radius=28, fg_color=GRAY_COLOR, width=400, height=350, bg_color=self._WHITE_COLOR)
        frame.grid(row=1, column=0, pady=(10,0))
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_propagate(False)

        settings_text = CTk.CTkLabel(master=frame, text=self._SETTINGS_TITLE_TEXT, font=(self._FONT,30),fg_color=GRAY_COLOR,text_color="black")
        settings_text.grid(row=0, column=0, pady=(10,0))

        choose_directory_text = CTk.CTkLabel(master=frame, text=self._SETTINGS_SELECT_FOLDER_TEXT, font=(self._FONT,18),fg_color="#c9c9c9",text_color="black")
        choose_directory_text.grid(row=1, column=0, pady=(10,0))

        select_folder_frame = CTk.CTkFrame(master=frame, fg_color=GRAY_COLOR)
        select_folder_frame.grid(row=2, column=0, pady=(5,0))

        self.selected_folder_entry = CTk.CTkEntry(master=select_folder_frame, text_color="black",fg_color=self._WHITE_COLOR, corner_radius=5, font=(self._FONT,10),width=250)
        
        if self._app.config_manager.save_folder_path:
            self.selected_folder_entry.insert(0, self._app.config_manager.save_folder_path)
        else:
            self.selected_folder_entry.insert(0, Path.cwd())

        self.selected_folder_entry.configure(state="disabled")
        self.selected_folder_entry.grid(row=0,column=0)

        select_folder_button = CTk.CTkButton(master=select_folder_frame, command=self._app.on_click_select_folder, image=self._folder_image, text="", width=20, border_width=1, border_color="black")
        select_folder_button.grid(row=0, column=1, padx=(3,0))

        clear_folder_button = CTk.CTkButton(master=select_folder_frame, command=self._app.on_click_remove_folder, image=self._delete_image, text="", width=20, border_width=1, border_color="black")
        clear_folder_button.grid(row=0, column=2, padx=(3,0))

        delete_templates_text = CTk.CTkLabel(master=frame, text=self._SETTINGS_DELETE_TEMPLATES_TEXT, font=(self._FONT,24),fg_color="#c9c9c9",text_color="black")
        delete_templates_text.grid(row=3, column=0, pady=(30,0))

        if len(self._app.config_manager.templates) != 0:
            templates_frame = CTk.CTkFrame(master=frame, fg_color=GRAY_COLOR)
            templates_frame.grid(row=4,column=0)

            for n,i in enumerate(self._app.config_manager.templates):
                delete_template_button = CTk.CTkButton(master=templates_frame,command=lambda name=i:(self._app.on_click_delete_template(name), delete_template_button.destroy()),hover_color=self._HOVER_PURPLE_COLOR,width=80,height=20, text_color="black",corner_radius=11,border_width=1,border_color="black",text=i,font=(self._FONT,18), fg_color=self._PURPLE_COLOR, bg_color=GRAY_COLOR)
                delete_template_button.grid(row=0, column=n, padx=5)
        else:
            not_found_text = CTk.CTkLabel(master=frame, text="Not found", font=(self._FONT,20),fg_color="#c9c9c9",text_color="black")
            not_found_text.grid(row=4,column=0, pady=(10,0))

        back_button = CTk.CTkButton(master=self,command=lambda:self._app.change_window(0),bg_color=self._WHITE_COLOR,width=100, text=self._BACK_BUTTON_TEXT, font=(self._FONT,24), fg_color=self._PURPLE_COLOR,hover_color=self._HOVER_PURPLE_COLOR, border_width=1, border_color="black",text_color="black")
        back_button.grid(row=2, column=0, padx=(0,110), pady=(10,0))

        save_button = CTk.CTkButton(master=self,command=self._app.on_click_save_settings,bg_color=self._WHITE_COLOR,width=100, text=self._SAVE_BUTTON_TEXT, font=(self._FONT,24), fg_color=self._PURPLE_COLOR,hover_color=self._HOVER_PURPLE_COLOR, border_width=1, border_color="black",text_color="black")
        save_button.grid(row=2, column=0, padx=(110,0), pady=(10,0))

    def load_add_tamplate_modal_window(self):
        background_frame = CTk.CTkFrame(master=self, width=800, height=427, fg_color=self._WHITE_COLOR,bg_color=self._WHITE_COLOR)
        background_frame.grid_propagate(False)
        background_frame.grid_columnconfigure(0, weight=1)
        background_frame.grid(row=1, column=0)

        frame = CTk.CTkFrame(master=background_frame, width=300, height=180, corner_radius=20, border_color="black", border_width=1, fg_color="#dedede")
        frame.grid_propagate(False)
        frame.grid_columnconfigure(0, weight=1)
        frame.grid(row=0, column=0, pady=(100,10))

        title = CTk.CTkLabel(master=frame, fg_color="#dedede", text_color="black", text=self._MAIN_MODAL_WINDOW_TITLE, font=(self._FONT, 24))
        title.grid(row=0, column=0, pady=(30,10))

        entry = CTk.CTkEntry(master=frame,width=180, text_color="black",fg_color=self._WHITE_COLOR, font=(self._FONT,20))
        entry.grid(row=1,column=0, pady=(0,10))

        back_button = CTk.CTkButton(master=frame,command=lambda:self._app.change_window(1),bg_color=self._WHITE_COLOR,width=90, text=self._BACK_BUTTON_TEXT, font=(self._FONT,22), fg_color=self._PURPLE_COLOR,hover_color=self._HOVER_PURPLE_COLOR, border_width=1, border_color="black",text_color="black")
        back_button.grid(row=2, column=0,padx=(0,100))
        save_button = CTk.CTkButton(master=frame,command=lambda:(self._app.save_tamplate(entry.get()),self._app.change_window(1)),bg_color=self._WHITE_COLOR,width=90, text=self._SAVE_BUTTON_TEXT, font=(self._FONT,22), fg_color=self._PURPLE_COLOR,hover_color=self._HOVER_PURPLE_COLOR, border_width=1, border_color="black",text_color="black")
        save_button.grid(row=2, column=0,padx=(100,0))

    #Clear window
    def _clear(self):
        for e in self.winfo_children():
            e.destroy()
    
    def _load_header(self, is_show_settings=False):
        header_frame = CTk.CTkFrame(master=self,border_width=0, fg_color=self._BLUE_COLOR, width=800, height=73, corner_radius=0)
        header_frame__text = CTk.CTkLabel(master=self, text=self._HEADER_TEXT, bg_color=self._BLUE_COLOR,font=(self._FONT,40))
        header_frame__text.grid(row=0, column=0)
        if(is_show_settings == True):
            header_settings = CTk.CTkButton(master=self, command=self.load_settings,image=self._settings_image, hover_color=self._BLUE_COLOR, fg_color=self._BLUE_COLOR,bg_color=self._BLUE_COLOR, text="", width=50)
            header_settings.grid(row=0,column=0, sticky="e",padx=20)
        header_frame.grid(row=0, column=0)

#Class that combines all classes and connects ui with logic
class App:
    def __init__(self):
        print("Starting application...")
        self.config_manager = None
        self.selected_fields = None
        self.html_parser = None
        self.filename = ""
        self.gui = Gui(self)
        self._read_config()
        self.gui.mainloop()

    #Read config and save it on config manager
    def _read_config(self):
        try:
            #Read config file
            self._file_manager = FileManager()
            #Save config on config manager
            self.config_manager = ConfigManager(self._file_manager.config)
        #Error when config file not found
        except FileNotFoundError:
            self.gui.error_text.configure(text="Config file not found")
        #Error when config has some syntaxes problem
        except json.JSONDecodeError:
            self.gui.error_text.configure(text="Cant read config file")

    #Parse html
    def _parse_html(self):
        self.html_parser = HtmlParcer(self._html)
    #Write table to excel file
    def _write_to_excel(self):
        
        #We check if it is possible to take data from the input for the file name, 
        #because if you use a custom order, then the input is not rendered and there will be an error, 
        #and if an error occurs, then write the last state of the input

        try:
            total_lines = self.gui.column_count_frame_entry.get()
        except:
            total_lines = self.total_lines

        #Checking if it possible to convert the data from the input into a number, 
        #if it is not possible, then the input is invalid and we put 0
        try:
            
            total_lines = int(total_lines)
        except:
            total_lines = 0

        try:
            filename = self.gui.file_name_frame_entry.get()
        except:
            filename = self.filename
        #if the input for the file name is empty, then we set the basic file name students
        if len(filename) == 0:
            filename = "students"
        self.excel_writer = ExcelWriter(self._data_manager.table,total_lines , self.config_manager.save_folder_path, filename)

    #Handle checkbox, when chekbox is pressed makes the input active or inactive
    def on_select_checkbox(self, *args):
            #Get index from args, -1 because we need index 
            index = int(args[3])-1
            #Current checkbox
            element = self.gui.checkboxes_list[index]
            #Get value from checkbox, 1 or 0
            value = element.get()
            if value == 1:
                #If 1, make same line input active
                self.gui.entry_list[index].configure(state="normal", fg_color=self.gui._WHITE_COLOR)
            else:
                #If 0, clear input and make inactive
                self.gui.entry_list[index].delete(0, CTk.END)
                self.gui.entry_list[index].configure(state="disabled", fg_color="#D5D5D5")

    #Set the same state for all checkboxes
    def select_all_checkboxes(self, value):
        for n in range(0, len(self.gui.checkboxes_list)):
            if value == 1:
                #Select checkbox
                self.gui.checkboxes_list[n].select()
                #Make input active
                self.gui.entry_list[n].configure(state="normal", fg_color=self.gui._WHITE_COLOR)
            else:
                #Deselect checkbox
                self.gui.checkboxes_list[n].deselect()
                #Clear input
                self.gui.entry_list[n].delete(0, CTk.END)
                #Make input inactive
                self.gui.entry_list[n].configure(state="disabled", fg_color="#D5D5D5")

    #Start data processing
    def compilate_data(self):
        #Get custom checkbox order status
        custom_order = self.gui.is_custom_order.get()
        self._data_manager = DataManager(self)

        #If no field is selected simply stop the function
        if len(self._data_manager.selected_fields) == 0:
            return

        #If the custom order is not selected,
        #then process the data, save it in Excel and load the final window
        if custom_order == 0:
            self._data_manager.process_data()
            self._write_to_excel()
            self.change_window(3)
        #If a custom order is selected, the program will open a window for custom order
        else:
            #Save the current text of the input file name
            self.filename = self.gui.file_name_frame_entry.get()
            self.total_lines = self.gui.column_count_frame_entry.get()
            self.selected_fields = self._data_manager.selected_fields
            self.change_window(2)

    #Handler for pressing the Paste button in the menu
    def menu_button_handle(self):
        #Ifno configuration, then there will simply be an error text and the program will not continue
        if self.config_manager == None: return

        #Try to get data from clipboard
        try:
            self._html = self.gui.clipboard_get()
        except:
            #
            self.gui.error_text.configure(text=self.gui._MENU_ERR_CANT_GET_CLIPBOARD_TEXT)
            return

        try:
            self._parse_html()
        except:
            self.gui.error_text.configure(text=self.gui._MENU_ERR_WRONG_HTML_CODE_TEXT)
            return
        
        #If all the checks were successful, load the main window
        self.gui.load_main()

    #Select checkboxes and fills inputs using a template
    def select_checkboxes_by_template(self, name):
        #Get template by name
        template = self.config_manager.templates[name]
        #Deselect all checkboxes
        self.select_all_checkboxes(0)
        #Iterate template (template is array)
        for i in template:
            #Iterate all checkboxes
            for n2,j in enumerate(self.gui.checkboxes_list):
                #Get text from checkbox (checkbox text and field name are same)
                text = j.cget("text")
                #Checking if field name is the same as text
                if text == i[0]:
                    #If yes, select checkbox
                    j.select()
                    #Make input active
                    self.gui.entry_list[n2].configure(state="normal", fg_color=self.gui._WHITE_COLOR)
                    #If element (array) lenght is 2 its mean that second element is preferred name
                    if len(i) == 2:
                        self.gui.entry_list[n2].delete(0, CTk.END)
                        self.gui.entry_list[n2].insert(0, i[1])

    #Pressing enter switches the focus to the next active input
    def focus_next_entry(self, event):
        #Get current input
        current_widget = event.widget
        #Find current input index
        for i, ctk_entry in enumerate(self.gui.entry_list):
            #If this input current
            if ctk_entry._entry == current_widget:
                #Minimum possible index 
                print(i)
                next_index = i + 1
                #Find next acrive index and move cursor
                while next_index < len(self.gui.entry_list):
                    #get next input by index
                    next_widget = self.gui.entry_list[next_index]
                    #Check if next input is active
                    if next_widget.cget("state") == "normal":
                        #Change focus
                        next_widget.focus_set()
                        #Stop loop
                        break
                    #if next input is disabled add +1 to next index and repeat
                    next_index += 1
                break
    #Return list for drag and drop system
    def get_select_fields_for_drag(self):
        #Refactor selected_fields [["chemistry",che"],["physics","phy"],["Math"]] -> ["che","phy","Math"]
        list = [field[-1] for field in self.selected_fields]
        #Delete first element because student names always first.
        del list[0]
        #Return refactored list
        return list
    
    #
    def compilate_ordered_data(self, new_order):
        new_order = ("Opiskelijan nimi",) + new_order
        sorted_fields = sorted(self.selected_fields, key=lambda x: new_order.index(x[-1]))
        self._data_manager.process_data(sorted_fields)
        self._write_to_excel()
        self.change_window(3)

    #Load window by index
    def change_window(self, index):
        if index == 0:
            self.gui.load_menu()
        elif index == 1:
            self.gui.load_main()
        elif index == 2:
            self.gui.load_custom_order()
        elif index == 3:
            self.gui.load_success()

    #
    def on_click_select_folder(self):
        #Select directory by modal window and save it
        directory = tk.filedialog.askdirectory()
        if(directory):
            #Make input active
            self.gui.selected_folder_entry.configure(state="normal")
            #Clear input
            self.gui.selected_folder_entry.delete(0, CTk.END)
            #Paste directory path to input
            self.gui.selected_folder_entry.insert(0, directory)
            #Paste directory path to config
            self.config_manager.save_folder_path = directory
            #Disable input for user
            self.gui.selected_folder_entry.configure(state="disabled")
    #save settings to config file
    def on_click_save_settings(self):
        #Write to config file
        self._file_manager.write_config(self.config_manager)
        #Change window to menu
        self.change_window(0)
    
    #Delete selected folder for saving files, files will be saved in the same folder as the exe file
    def on_click_remove_folder(self):
        #Make input active
        self.gui.selected_folder_entry.configure(state="normal")
        #clear input
        self.gui.selected_folder_entry.delete(0, CTk.END)
        #Paste exe folder directory to input
        self.gui.selected_folder_entry.insert(0, Path.cwd())
        #Save empty path to config
        self.config_manager.save_folder_path = ""
        #make input disabled
        self.gui.selected_folder_entry.configure(state="disabled")

    #Delete template
    def on_click_delete_template(self, name):
        #Delete template from config
        del self.config_manager.templates[name]
        #Write new config to config file
        self._file_manager.write_config(self.config_manager)

    #Save new template to config and config file
    def save_tamplate(self, name):
        #Create an instance of the data manager class
        data_manager = DataManager(self)
        #Save new template to config
        self.config_manager.templates[name] = data_manager.selected_fields
        #Write new config to config file
        self._file_manager.write_config(self.config_manager)

    #Return relative path, 
    #this is necessary so that there are no errors 
    #when compiling the program and when assembling the program into one exe file
    def resource_path(self,relative_path):
        #Get path when is exe file
        try:
            base_path = sys._MEIPASS
        #Get path when run source
        except Exception:
            base_path = os.path.abspath(".")
        #Return path
        return os.path.join(base_path, relative_path)

if __name__ == "__main__":
    app = App()